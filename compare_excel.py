from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def compare_and_highlight_with_format(file1: str, file2: str, output_file: str = "diff_result.xlsx"):
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    # Copy file2 làm nền kết quả (giữ nguyên format)
    wb_result = load_workbook(file2)

    # Màu highlight
    fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # đỏ nhạt
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # xanh nhạt

    for sheet_name in wb_result.sheetnames:
        ws1 = wb1[sheet_name] if sheet_name in wb1.sheetnames else None
        ws2 = wb2[sheet_name]
        ws_result = wb_result[sheet_name]

        max_rows = max(ws1.max_row if ws1 else 0, ws2.max_row)
        max_cols = max(ws1.max_column if ws1 else 0, ws2.max_column)

        sheet_has_diff = False  # cờ để check nếu sheet có khác biệt

        for r in range(1, max_rows + 1):
            for c in range(1, max_cols + 1):
                val1 = ws1.cell(r, c).value if ws1 else None
                val2 = ws2.cell(r, c).value

                if val1 != val2:
                    sheet_has_diff = True
                    cell = ws_result.cell(r, c)

                    if val1 is not None and val2 is not None:
                        cell.fill = fill_green  # thay đổi giá trị -> highlight xanh
                    elif val1 is None and val2 is not None:
                        cell.fill = fill_green  # mới thêm ở file2
                    elif val1 is not None and val2 is None:
                        cell.fill = fill_red    # bị xóa ở file2

        # Nếu sheet có khác biệt -> đổi màu tab sheet (cam)
        if sheet_has_diff:
            ws_result.sheet_properties.tabColor = "FFC000"  # cam

    wb_result.save(output_file)
    print(f"✅ Đã lưu file {output_file}, highlight ô khác biệt và đổi màu tab sheet có khác biệt.")


# 👉 Gọi hàm
compare_and_highlight_with_format("D:\\Temp\\C2.3.2-8_画面仕様書_KYHK_G_01_受注管理.xlsx", "D:\\Temp\\C2.3.2-8_画面仕様書_KYHK_G_01_受注管理_new.xlsx", "D:\\Temp\\C2.3.2-8_画面仕様書_KYHK_G_01_受注管理_diff_result.xlsx")
